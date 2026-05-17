#!/usr/bin/env python3
import argparse
import json
import openpyxl
from datetime import datetime

def parse_date(date_str):
    if not date_str:
        return None
    if isinstance(date_str, str):
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            return None
    return date_str

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--drop_id', required=True)
    parser.add_argument('--today', required=True)
    parser.add_argument('--output_json', required=True)
    args = parser.parse_args()

    today = datetime.strptime(args.today, '%Y-%m-%d').date()
    wb = openpyxl.load_workbook(args.workbook)
    
    # Get drop
    drop_sheet = wb['Drop Plan']
    drop_data = None
    drop_row = None
    
    for idx, row in enumerate(drop_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == args.drop_id:
            drop_data = row
            drop_row = idx
            break
    
    if not drop_data:
        print(f"Drop {args.drop_id} not found")
        return
    
    # Get Document Library
    doc_sheet = wb['Document Library']
    docs = {}
    for row in doc_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            docs[row[0]] = {'title': row[3], 'version': row[4], 'path': row[6]}
    
    # Get Employees
    emp_sheet = wb['Employees']
    employees = {}
    for row in emp_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            employees[row[0]] = {'name': row[1], 'trade': row[2]}
    
    # Parse drop data
    headers = [cell.value for cell in drop_sheet[1]]
    drop_dict = dict(zip(headers, drop_data))
    
    planned_start = parse_date(drop_dict['Planned Start Date'])
    planned_end = parse_date(drop_dict['Planned End Date'])
    actual_start = parse_date(drop_dict['Actual Start Date'])
    actual_end = parse_date(drop_dict['Actual End Date'])
    
    # Calculate derived fields
    days_remaining = (planned_end - today).days if planned_end else None
    if planned_start and planned_end:
        total_days = (planned_end - planned_start).days + 1
        elapsed = (today - planned_start).days + 1
        pct_complete = min(100, max(0, int((elapsed / total_days) * 100)))
    else:
        pct_complete = 0
    
    is_critical = days_remaining is not None and days_remaining < 0 and drop_dict['Status'] != 'Signed Off'
    
    # Sign-off progress
    sign_off_required = [s.strip() for s in str(drop_dict['Sign-Off Required From']).split(';') if s.strip()]
    sign_off_received = []
    if drop_dict['Sign-Off Foreman']:
        sign_off_received.append('Foreman')
    if drop_dict['Sign-Off Superintendent']:
        sign_off_received.append('Superintendent')
    if drop_dict['Sign-Off QEI']:
        sign_off_received.append('QEI')
    if drop_dict['Sign-Off Owner Rep']:
        sign_off_received.append('Owner Rep')
    
    sign_off_pending = [s for s in sign_off_required if s not in sign_off_received]
    
    # Crew names
    crew_ids = [c.strip() for c in str(drop_dict['Crew Assigned']).split(';') if c.strip() and c.strip() != 'TBD']
    crew_members = []
    for crew_id in crew_ids:
        if crew_id in employees:
            crew_members.append({
                'id': crew_id,
                'name': employees[crew_id]['name'],
                'trade': employees[crew_id]['trade']
            })
    
    # Drawing refs
    drawing_refs = []
    drawing_ids = [d.strip() for d in str(drop_dict['Drawing References']).split(';') if d.strip()]
    for doc_id in drawing_ids:
        if doc_id in docs:
            drawing_refs.append({
                'id': doc_id,
                'title': docs[doc_id]['title'],
                'version': docs[doc_id]['version'],
                'path': docs[doc_id]['path']
            })
    
    # Successors and predecessors
    successors_str = str(drop_dict.get('Successor Drops', '')).strip()
    successors = [s.strip() for s in successors_str.split(';') if s.strip()]
    
    predecessors_str = str(drop_dict.get('Predecessor Drops', '')).strip()
    predecessors = [p.strip() for p in predecessors_str.split(';') if p.strip()]
    
    # Next/previous drop
    next_drop = successors[0] if successors else None
    previous_drop = predecessors[0] if predecessors else None
    
    output = {
        'drop_id': drop_dict['Drop ID'],
        'project_code': drop_dict['Project Code'],
        'elevation': drop_dict['Elevation'],
        'bay_range': drop_dict['Bay Range'],
        'floor_range': drop_dict['Floor Range'],
        'scope_of_work': drop_dict['Scope of Work'],
        'trades_required': [t.strip() for t in str(drop_dict['Trade(s) Required']).split(';') if t.strip()],
        'estimated_duration_days': drop_dict['Estimated Duration (Days)'],
        'planned_start': str(planned_start) if planned_start else None,
        'planned_end': str(planned_end) if planned_end else None,
        'actual_start': str(actual_start) if actual_start else None,
        'actual_end': str(actual_end) if actual_end else None,
        'crew_size': drop_dict['Crew Size'],
        'crew_members': crew_members,
        'equipment_required': [e.strip() for e in str(drop_dict['Equipment Required']).split(';') if e.strip()],
        'materials_required': [m.strip() for m in str(drop_dict['Materials Required']).split(';') if m.strip()],
        'drawing_references': drawing_refs,
        'status': drop_dict['Status'],
        'sign_off_required': sign_off_required,
        'sign_off_status': drop_dict['Sign-Off Status'],
        'sign_off_dates': {
            'foreman': str(drop_dict['Sign-Off Foreman']) if drop_dict['Sign-Off Foreman'] else None,
            'superintendent': str(drop_dict['Sign-Off Superintendent']) if drop_dict['Sign-Off Superintendent'] else None,
            'qei': str(drop_dict['Sign-Off QEI']) if drop_dict['Sign-Off QEI'] else None,
            'owner_rep': str(drop_dict['Sign-Off Owner Rep']) if drop_dict['Sign-Off Owner Rep'] else None,
        },
        'photos_required': drop_dict['Photos Required'],
        'photos_captured': drop_dict['Photos Captured'],
        'notes': drop_dict['Notes'],
        'days_remaining': days_remaining,
        'pct_complete': pct_complete,
        'is_critical': is_critical,
        'sign_off_progress': {
            'required': sign_off_required,
            'received': sign_off_received,
            'pending': sign_off_pending
        },
        'predecessor_drops': predecessors,
        'successor_drops': successors,
        'next_drop': next_drop,
        'previous_drop': previous_drop,
    }
    
    with open(args.output_json, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"Generated {args.output_json}")

if __name__ == '__main__':
    main()
