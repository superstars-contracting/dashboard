import openpyxl
import json
import argparse
from datetime import datetime, date

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return str(obj)
        return super().default(obj)

def generate_rfi(workbook_path, rfi_number, today, output_json):
    wb = openpyxl.load_workbook(workbook_path)
    
    if 'RFI Log' not in wb.sheetnames:
        print(f"Error: RFI Log sheet not found")
        return

    ws = wb['RFI Log']
    emp_ws = wb['Employees'] if 'Employees' in wb.sheetnames else None
    
    rfi_row = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == rfi_number:
            rfi_row = row
            break

    if not rfi_row:
        print(f"Error: {rfi_number} not found in RFI Log")
        return

    emp_data = {}
    if emp_ws:
        for row in emp_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                emp_data[row[0]] = {
                    'name': row[1],
                    'position': row[2],
                    'company_trade': row[3] if len(row) > 3 else 'Superstar Contracting',
                    'phone': row[6] if len(row) > 6 else '',
                    'email': row[7] if len(row) > 7 else ''
                }

    [rfi_id, proj_code, date_sub, emp_id, emp_name, company, title, desc, category, location,
     q, sch, cost, qual, safe, stop, priority, dist, status, due_date, resp_date, resp_summary, photo_count, ref_docs] = rfi_row

    date_sub = str(date_sub) if date_sub else None
    due_date = str(due_date) if due_date else None
    resp_date = str(resp_date) if resp_date else None

    impacts = []
    if sch == 'Y':
        impacts.append('Schedule delay')
    if cost == 'Y':
        impacts.append('Cost impact')
    if qual == 'Y':
        impacts.append('Quality risk')
    if safe == 'Y':
        impacts.append('Safety issue')
    if stop == 'Y':
        impacts.append('Work stoppage')

    distribution = [d.strip() for d in (dist or '').split(',')] if dist else []

    days_overdue = None
    if status in ['Submitted', 'Under Review', 'Overdue'] and due_date:
        try:
            due = datetime.fromisoformat(due_date)
            today_dt = datetime.fromisoformat(today)
            if today_dt > due:
                days_overdue = (today_dt - due).days
        except:
            pass

    submitter_info = emp_data.get(emp_id, {})

    payload = {
        'rfi_id': rfi_id,
        'project': {
            'code': proj_code,
            'name': '890 E 135th St — Mott Haven Restoration',
            'address': '890 East 135th Street, Bronx, NY 10454'
        },
        'submitter': {
            'employee_id': emp_id,
            'name': emp_name,
            'company_trade': company,
            'phone': submitter_info.get('phone', ''),
            'email': submitter_info.get('email', '')
        },
        'details': {
            'title': title,
            'description': desc,
            'category': category,
            'location': location,
            'date_submitted': date_sub
        },
        'question': {
            'primary': q,
            'impacts': impacts
        },
        'priority': priority,
        'distribution_list': distribution,
        'status': status,
        'response': {
            'due_date': due_date,
            'received_date': resp_date,
            'summary': resp_summary,
            'days_overdue': days_overdue
        },
        'attachments': {
            'photo_count': int(photo_count) if photo_count else 0,
            'reference_documents': [d.strip() for d in (ref_docs or '').split(',')] if ref_docs else []
        },
        'metadata': {
            'generated_at': today,
            'source_workbook': workbook_path,
            'warnings': []
        }
    }

    with open(output_json, 'w') as f:
        json.dump(payload, f, indent=2, cls=DateTimeEncoder)

    print(f"Generated {output_json}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--rfi_number', required=True)
    parser.add_argument('--today', required=True)
    parser.add_argument('--output_json', required=True)
    args = parser.parse_args()

    generate_rfi(args.workbook, args.rfi_number, args.today, args.output_json)
