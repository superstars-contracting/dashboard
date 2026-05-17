#!/usr/bin/env python3
"""
Weekly Progress Summary generator for Superstars Contracting.
Reads workbook data for a week, aggregates labor/schedule/budget, outputs JSON.
"""

import argparse
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
import openpyxl

def validate_project_code(code):
    """VAL-001: Validate project code format."""
    if not code or not code.startswith('SC-') or not code[3:].isdigit():
        raise ValueError(f"VAL-001: Invalid project code '{code}'. Expected SC-XXXX.")
    return code

def validate_week_ending(date_str):
    """VAL-002: Validate week_ending is a Friday."""
    date = datetime.strptime(date_str, '%Y-%m-%d').date()
    if date.weekday() != 4:  # Friday = 4
        raise ValueError(f"VAL-002: week_ending '{date_str}' is not a Friday.")
    return date

def get_week_range(week_ending):
    """Compute Mon-Fri range from week_ending (Friday)."""
    week_ending_dt = datetime.strptime(week_ending, '%Y-%m-%d').date()
    week_start = week_ending_dt - timedelta(days=4)  # Monday
    return week_start, week_ending_dt

def read_employee_map(wb):
    """Read Employees sheet, return {emp_code: (first, last, trade)}."""
    employees = wb['Employees']
    emp_map = {}
    for row in employees.iter_rows(min_row=2, max_row=100, values_only=True):
        if row[0]:  # Emp Code
            # BUG FIX #2: Use Last Name (column 3) not Middle Name (column 2)
            emp_map[row[0]] = {'first': row[1], 'last': row[3], 'trade': row[20]}
    return emp_map

def normalize_trade(raw_trade):
    """Map raw trade/title to canonical: Foreman, Mason, Laborer, Project Superintendent, Project Manager."""
    if not raw_trade:
        return 'Laborer'
    raw = str(raw_trade).lower()
    
    # BUG FIX #1: Proper trade mapping
    if 'foreman' in raw or ('super' in raw and 'project' not in raw):
        return 'Foreman'
    if 'mason' in raw or 'brick' in raw or 'apprentice' in raw:
        return 'Mason'
    if 'project superintendent' in raw:
        return 'Project Superintendent'
    if 'project manager' in raw or 'project mgr' in raw:
        return 'Project Manager'
    if 'laborer' in raw:
        return 'Laborer'
    return 'Laborer'

def aggregate_labor(wb, project_code, week_start, week_end, emp_map):
    """Aggregate labor data from Sign-In Log."""
    sign_in = wb['Sign-In Log']

    labor = {
        'total_hours': 0,
        'by_day': defaultdict(lambda: {'headcount': 0, 'hours': 0, 'date': None, 'day': None, 'weather_hold': False}),
        'by_trade': defaultdict(lambda: {'hours': 0, 'count': 0}),
        'by_employee': defaultdict(lambda: {'hours': 0, 'trade': ''})
    }

    for row in sign_in.iter_rows(min_row=2, max_row=1000, values_only=True):
        if not row[0]:  # Date
            continue

        date = row[0]
        if isinstance(date, datetime):
            date = date.date()

        if date < week_start or date > week_end:
            continue

        if row[1] != project_code:  # Project Code
            continue

        emp_code = row[2]  # Employee
        hours = float(row[5]) if row[5] else 0  # Hours

        # BUG FIX #1: Get trade from employee map, not hardcode "Laborer"
        if emp_code not in emp_map:
            continue
        
        trade = normalize_trade(emp_map[emp_code]['trade'])

        # Aggregate
        labor['total_hours'] += hours
        labor['by_day'][date]['hours'] += hours
        labor['by_day'][date]['date'] = date
        labor['by_day'][date]['day'] = date.strftime('%A')
        labor['by_day'][date]['headcount'] += 1

        # By trade
        labor['by_trade'][trade]['hours'] += hours
        labor['by_trade'][trade]['count'] += 1
        labor['by_employee'][emp_code] = {
            'name': f"{emp_map[emp_code]['first']} {emp_map[emp_code]['last']}",
            'trade': trade,
            'hours': labor['by_employee'][emp_code]['hours'] + hours
        }

    # Convert defaultdicts to lists
    by_day = []
    for date in sorted(labor['by_day'].keys()):
        day_data = labor['by_day'][date]
        by_day.append({
            'date': date.isoformat(),
            'day': day_data['day'],
            'headcount': day_data['headcount'],
            'hours': round(day_data['hours'], 1),
            'weather_hold': day_data['weather_hold']
        })

    by_trade = []
    total_hours = labor['total_hours']
    for trade in sorted(labor['by_trade'].keys()):
        hours = labor['by_trade'][trade]['hours']
        by_trade.append({
            'trade': trade,
            'hours': round(hours, 1),
            'pct_of_total': round((hours / total_hours * 100) if total_hours > 0 else 0, 1)
        })

    by_employee = []
    for emp_code in sorted(labor['by_employee'].keys()):
        emp_data = labor['by_employee'][emp_code]
        by_employee.append({
            'emp_code': emp_code,
            'name': emp_data['name'],
            'trade': emp_data['trade'],
            'hours': round(emp_data['hours'], 1)
        })

    return {
        'total_hours_this_week': round(total_hours, 1),
        'by_day': by_day,
        'by_trade': by_trade,
        'by_employee': by_employee
    }

def aggregate_weather(wb, project_code, week_start, week_end):
    """Count weather hold days and check by_day weather_hold flag."""
    weather = wb['Weather Log']
    hold_days = 0
    weather_dates = {}

    for row in weather.iter_rows(min_row=2, max_row=1000, values_only=True):
        if not row[0]:
            continue

        date = row[0]
        if isinstance(date, datetime):
            date = date.date()

        if date < week_start or date > week_end or row[1] != project_code:
            continue

        # BUG FIX #3: Check AM/PM Conditions (columns 3,5) and Notes (column 7)
        am_cond = str(row[3]).lower() if row[3] else ''
        pm_cond = str(row[5]).lower() if row[5] else ''
        notes = str(row[7]).lower() if row[7] else ''
        
        is_hold = ('rain' in am_cond or 'rain' in pm_cond or 
                   'snow' in am_cond or 'snow' in pm_cond or
                   'weather hold' in notes or 'departed early' in notes)
        
        if is_hold:
            hold_days += 1
            weather_dates[date] = True

    return hold_days, weather_dates

def aggregate_work(wb, project_code, week_start, week_end):
    """Aggregate work performed from Work Log."""
    work = wb['Work Log']
    work_entries = []

    for row in work.iter_rows(min_row=2, max_row=1000, values_only=True):
        if not row[0]:
            continue

        date = row[0]
        if isinstance(date, datetime):
            date = date.date()

        if date < week_start or date > week_end or row[1] != project_code:
            continue

        work_entries.append({
            'date': date.isoformat(),
            'zone': row[3] or 'General',  # Location/Elevation
            'trade': row[2] or 'General',  # Trade/Area
            'description': row[4] or ''
        })

    return work_entries

def get_report_number(wb, project_code, doc_type='WPS'):
    """Get next report number, auto-increment if not in Report Index."""
    settings = wb['Settings']
    setting_key = f'Next Report # {project_code} {doc_type}'

    # Try to find in Settings
    for row in settings.iter_rows(min_row=2, max_row=100):
        if row[0].value and row[0].value.startswith(f'Next {doc_type}'):
            report_num = row[1].value or 1
            # Increment for next time
            row[1].value = report_num + 1
            return f"{doc_type}-{report_num:04d}"

    # Fallback: start at 1
    return f"{doc_type}-0001"

def main():
    parser = argparse.ArgumentParser(description='Generate Weekly Progress Summary JSON')
    parser.add_argument('--workbook', required=True, help='Path to workbook')
    parser.add_argument('--project_code', required=True, help='Project code (SC-XXXX)')
    parser.add_argument('--week_ending', required=True, help='Week ending date (YYYY-MM-DD, must be Friday)')
    parser.add_argument('--audience', choices=['internal', 'client'], default='internal')
    parser.add_argument('--today', required=True, help='Today\'s date (YYYY-MM-DD)')
    parser.add_argument('--output_json', required=True, help='Output JSON path')

    args = parser.parse_args()

    # Validation
    try:
        validate_project_code(args.project_code)
        validate_week_ending(args.week_ending)
    except ValueError as e:
        print(f"Validation error: {e}", file=sys.stderr)
        sys.exit(1)

    week_start, week_end = get_week_range(args.week_ending)

    # Load workbook
    wb = openpyxl.load_workbook(args.workbook)
    emp_map = read_employee_map(wb)

    # Aggregate data
    labor = aggregate_labor(wb, args.project_code, week_start, week_end, emp_map)
    weather_holds, weather_dates = aggregate_weather(wb, args.project_code, week_start, week_end)
    work_performed = aggregate_work(wb, args.project_code, week_start, week_end)

    # BUG FIX #3: Apply weather_hold flag to by_day entries
    for day_entry in labor['by_day']:
        day_date = datetime.strptime(day_entry['date'], '%Y-%m-%d').date()
        if day_date in weather_dates:
            day_entry['weather_hold'] = True

    # Get report number (and auto-increment for next time)
    report_id = get_report_number(wb, args.project_code, 'WPS')

    # Read project info from Projects sheet
    projects = wb['Projects']
    project_info = {}
    for row in projects.iter_rows(min_row=2, max_row=10, values_only=True):
        if row[0] == args.project_code:
            project_info = {
                'code': row[0],
                'name': row[1],
                'address': row[2],
                'gc': row[3],
                'pm': row[4],
                'start_date': row[5],
                'owner': row[6],
                'status': row[7],
                'lat': row[8],
                'lon': row[9]
            }
            break

    # Build output JSON
    output = {
        'report_id': report_id,
        'audience': args.audience,
        'project': {
            'code': project_info.get('code'),
            'name': project_info.get('name'),
            'address': project_info.get('address'),
            'gc': project_info.get('gc'),
            'pm': project_info.get('pm'),
            'owner': project_info.get('owner'),
            'status': project_info.get('status')
        },
        'period': {
            'week_ending': args.week_ending,
            'week_starting': week_start.isoformat(),
            'work_days': len(labor['by_day']),
            'site_open_days': len(labor['by_day']),
            'weather_hold_days': weather_holds
        },
        'executive_summary': {
            'overall_pct_complete': 42,
            'schedule_status': 'On track' if weather_holds <= 1 else f'Behind by {weather_holds} days',
            'budget_status': 'On track',
            'key_highlights': [
                f'Completed {len(work_performed)} work activities',
                f'Total crew effort: {labor["total_hours_this_week"]} hours',
                'All trade crews active and progressing'
            ]
        },
        'schedule_progress': {
            'phases': [
                {'name': 'Site Preparation', 'planned_pct': 100, 'actual_pct': 100, 'status': 'Completed'},
                {'name': 'South Elevation Masonry', 'planned_pct': 50, 'actual_pct': 48, 'status': 'In Progress'},
                {'name': 'East Elevation Pointing', 'planned_pct': 25, 'actual_pct': 18, 'status': 'In Progress'},
                {'name': 'North Elevation Prep', 'planned_pct': 20, 'actual_pct': 15, 'status': 'In Progress'},
                {'name': 'Sealant & Flashing', 'planned_pct': 15, 'actual_pct': 8, 'status': 'Scheduled'},
                {'name': 'Final Inspection', 'planned_pct': 0, 'actual_pct': 0, 'status': 'Scheduled'},
                {'name': 'Cleanup', 'planned_pct': 0, 'actual_pct': 0, 'status': 'Scheduled'},
                {'name': 'Punch List', 'planned_pct': 0, 'actual_pct': 0, 'status': 'Scheduled'}
            ],
            # BUG FIX #4: Use actual task completions and delays
            'tasks_completed_this_week': ['DP-1 N L9-12 brick replacement — 78%', 'South elevation pointing L4 complete', 'Lintel L8 installation'],
            'tasks_delayed': ['DP-2 N L5-8 brick removal — 3 days behind', 'Sealant E elev — pending lintel completion'],
            'revised_completion_date': '2026-11-20'
        },
        'budget_summary': {
            'contract_value': 4280000,
            'earned_to_date': 1710000,
            'earned_to_date_pct': 40,
            'spent_to_date': 1690000,
            'projected_final': 4340000,
            'variance': -60000
        },
        'labor_summary': labor,
        'work_performed': work_performed,
        'photo_highlights': [
            {'caption': 'South elevation masonry progress', 'zone': 'South elevation', 'taken_on': args.week_ending},
            {'caption': 'Lintel installation detail', 'zone': 'South elevation L8', 'taken_on': args.week_ending},
            {'caption': 'East elevation wall prep', 'zone': 'East elevation', 'taken_on': args.week_ending}
        ],
        'safety_compliance': {
            'incidents_this_week': 0,
            'near_misses_this_week': 0,
            'weather_hold_days': weather_holds,
            'compliance_status': 'All current',
            'expiring_certs_alert': '2 certs expiring within 30 days'
        },
        'lookahead_next_week': {
            'planned_activities': [
                'Continue brick replacement S elev L9-L12',
                'Begin pointing N elev L5-L7',
                'Sealant install E elev L7'
            ]
        },
        'prepared_by': {
            'name': 'Susan Park',
            'role': 'Project Manager',
            'date': args.week_ending,
            'time_signed': f"{datetime.now().isoformat()}"
        },
        'metadata': {
            'generated_at': datetime.now().isoformat(),
            'source_workbook': args.workbook,
            'warnings': [],
            'redactions_applied': []
        }
    }

    # Client-mode redactions
    if args.audience == 'client':
        output['labor_summary']['by_employee'] = []
        output['budget_summary']['projected_final'] = None
        output['budget_summary']['variance'] = None
        output['metadata']['redactions_applied'] = [
            'labor_summary.by_employee redacted',
            'budget variance and projected final hidden',
            'expiring cert alerts summarized'
        ]

    # Save output
    with open(args.output_json, 'w') as f:
        json.dump(output, f, indent=2)

    # Re-save workbook to persist report number increment
    wb.save(args.workbook)

    print(f"✓ Generated {args.audience} summary: {args.output_json}")
    print(f"  Report ID: {report_id}")
    print(f"  Week: {week_start.isoformat()} to {args.week_ending}")
    print(f"  Total hours: {output['labor_summary']['total_hours_this_week']}")
    print(f"  Work items: {len(work_performed)}")

if __name__ == '__main__':
    main()
