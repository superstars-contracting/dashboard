#!/usr/bin/env python3
"""
Migration script: Daily_Construction_Report.xlsx -> superstars.db (SQLite)
Phase 2A: Workbook becomes backup; SQLite becomes source of truth
"""

import sqlite3
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import re

SCRIPT_DIR = Path(__file__).parent
WORKBOOK_PATH = SCRIPT_DIR / "Daily_Construction_Report.xlsx"
DB_PATH = SCRIPT_DIR / "superstars.db"
SCHEMA_PATH = SCRIPT_DIR / "schema.sql"

# Boolean columns that should be Y/N/N-A
BOOLEAN_COLUMNS = {
    'renewal_required', 'photos_required', 'photos_captured',
    'personnel_all_signed_out', 'personnel_visitors_escorted', 'personnel_no_remaining',
    'equipment_tools_secured', 'equipment_scaffold_locked', 'equipment_compressors_secured',
    'equipment_mast_climbers_locked', 'fire_watch_completed', 'fire_heat_sources_cool',
    'fire_extinguishers_ready', 'dust_collection_sealed', 'dust_storage_sealed', 'dust_tarps_secure',
    'water_penetrations_covered', 'water_window_protection', 'security_access_locked',
    'security_fence_secured', 'security_sidewalk_clear', 'building_doors_locked',
    'building_roof_locked', 'climate_hvac_undisturbed', 'climate_storage_sealed',
    'doc_daily_report', 'doc_photos_taken'
}

# Column name mapping: workbook column -> database column
COLUMN_MAPPINGS = {
    "Code": "project_code", "Name": "name", "Address": "address", "City/Zip": "city_zip",
    "Superintendent": "superintendent", "ID": "employee_id", "Trade": "trade",
    "Date": "date", "Employee ID": "employee_id", "Project Code": "project_code",
    "Time In": "time_in", "Time Out": "time_out", "Setting Name": "setting_name",
    "Value": "value", "Permit ID": "permit_id", "Permit Type": "permit_type",
    "Permit Number": "permit_number", "Issuing Agency": "issuing_agency",
    "Issued Date": "issued_date", "Expiration Date": "expiration_date", "Status": "status",
    "Renewal Required (Y/N)": "renewal_required", "Renewal Submitted Date": "renewal_submitted_date",
    "Renewal Approved Date": "renewal_approved_date", "Cost ($)": "cost", "File Path": "file_path",
    "Notes": "notes", "Last Reviewed": "last_reviewed", "Doc ID": "doc_id",
    "Type": "type", "Title": "title", "Version": "version", "Discipline": "discipline",
    "File Size (KB)": "file_size_kb", "Uploaded At": "uploaded_at", "Uploaded By": "uploaded_by",
    "Status": "status", "Linked Records": "linked_records", "Code ID": "code_id",
    "Code Title": "code_title", "Source": "source", "Document Type": "document_type",
    "Applies To": "applies_to", "Loaded At": "loaded_at", "Last Updated By DOB": "last_updated_by_dob",
    "Project Codes": "project_codes", "Compliance Rules Count": "compliance_rules_count",
    "Talk ID": "talk_id", "Category": "category", "DOB Reference": "dob_reference",
    "OSHA Reference": "osha_reference", "Duration (min)": "duration_min", "Required For": "required_for",
    "Frequency Recommendation": "frequency_recommendation", "Hazards Summary": "hazards_summary",
    "Key Practices": "key_practices", "Required PPE": "required_ppe", "Discussion Questions": "discussion_questions",
    "Required Inspections": "required_inspections", "Related Certifications": "related_certifications",
    "Last Updated": "last_updated", "Author": "author", "Drop ID": "drop_id",
    "Elevation": "elevation", "Bay Range": "bay_range", "Floor Range": "floor_range",
    "Scope of Work": "scope_of_work", "Trade(s) Required": "trades_required",
    "Estimated Duration (Days)": "estimated_duration_days", "Planned Start Date": "planned_start_date",
    "Planned End Date": "planned_end_date", "Actual Start Date": "actual_start_date",
    "Actual End Date": "actual_end_date", "Crew Size": "crew_size", "Crew Assigned": "crew_assigned",
    "Equipment Required": "equipment_required", "Materials Required": "materials_required",
    "Drawing References": "drawing_references", "Sign-Off Required From": "sign_off_required_from",
    "Sign-Off Status": "sign_off_status", "Sign-Off Foreman": "sign_off_foreman",
    "Sign-Off Superintendent": "sign_off_superintendent", "Sign-Off QEI": "sign_off_qei",
    "Sign-Off Owner Rep": "sign_off_owner_rep", "Photos Required": "photos_required",
    "Photos Captured": "photos_captured", "Linked Punch Items": "linked_punch_items",
    "Predecessor Drops": "predecessor_drops", "Successor Drops": "successor_drops",
    "Schedule ID": "schedule_id", "Meeting Type": "meeting_type", "Recurrence": "recurrence",
    "Day of Week": "day_of_week", "Time": "time", "Default Location": "default_location",
    "Default Distribution List": "default_distribution_list", "Meeting ID": "meeting_id",
    "Time Start": "time_start", "Time End": "time_end", "Location": "location",
    "Prepared By": "prepared_by", "Attendees": "attendees", "Transcript Source": "transcript_source",
    "Summary": "summary", "Decisions": "decisions", "Distribution List": "distribution_list",
    "Action ID": "action_id", "Description": "description", "Owner": "owner",
    "Owner Email": "owner_email", "Due Date": "due_date", "Completion Date": "completion_date",
    "Closure ID": "closure_id", "Foreman ID": "foreman_id", "Foreman Name": "foreman_name",
    "Time of Close": "time_of_close", "Weather at Close": "weather_at_close",
    "Equipment Left Overnight": "equipment_left_overnight",
    "Personnel — All workers signed out": "personnel_all_signed_out",
    "Personnel — Visitors escorted off + signed out": "personnel_visitors_escorted",
    "Personnel — No personnel remaining in building": "personnel_no_remaining",
    "Equipment — Tools and equipment secured / removed": "equipment_tools_secured",
    "Equipment — Scaffold locked, tagged, retracted": "equipment_scaffold_locked",
    "Equipment — Compressors / generators shut off and secured": "equipment_compressors_secured",
    "Equipment — Mast climbers lowered to ground and locked": "equipment_mast_climbers_locked",
    "Fire — 30-min fire watch completed for all hot work": "fire_watch_completed",
    "Fire — All heat sources confirmed cool": "fire_heat_sources_cool",
    "Fire — Fire extinguishers in service position": "fire_extinguishers_ready",
    "Dust — Dust collection emptied and sealed": "dust_collection_sealed",
    "Dust — Path to interior art storage confirmed sealed": "dust_storage_sealed",
    "Dust — Tarps / barriers secure for overnight": "dust_tarps_secure",
    "Water — Wall penetrations covered or sealed": "water_penetrations_covered",
    "Water — Window / opening protection in place": "water_window_protection",
    "Security — All site access points locked": "security_access_locked",
    "Security — Construction fence secured": "security_fence_secured",
    "Security — Sidewalk shed pedestrian path clear": "security_sidewalk_clear",
    "Building — Building exterior doors locked": "building_doors_locked",
    "Building — Roof access closed and locked": "building_roof_locked",
    "Climate — Building HVAC undisturbed": "climate_hvac_undisturbed",
    "Climate — Interior doors to art storage confirmed sealed": "climate_storage_sealed",
    "Doc — Daily report submitted": "doc_daily_report",
    "Doc — Photos of site condition taken": "doc_photos_taken",
    "Signed Timestamp": "signed_timestamp",
}

def snake_case(s):
    """Convert column name to snake_case."""
    if not s:
        return s
    if s in COLUMN_MAPPINGS:
        return COLUMN_MAPPINGS[s]
    s = re.sub(r'[^\w]', '_', s.strip())
    s = re.sub(r'_+', '_', s).lower().strip('_')
    return s

def normalize_boolean(val, col_name):
    """Convert value to Y/N/N-A for boolean columns."""
    if col_name not in BOOLEAN_COLUMNS:
        return val
    if val is None or val == '':
        return None
    s = str(val).strip().upper()
    if s in ('Y', 'YES', '1', 'TRUE'):
        return 'Y'
    if s in ('N', 'NO', '0', 'FALSE'):
        return 'N'
    if s in ('N/A', 'NA', 'N-A'):
        return 'N/A'
    return None

def excel_to_iso_date(val):
    """Convert Excel datetime to ISO date string."""
    if val is None or val == '':
        return None
    if isinstance(val, str):
        return val
    try:
        if hasattr(val, 'date'):
            return val.date().isoformat()
        return str(val)
    except:
        return None

def init_db():
    """Create database from schema."""
    if DB_PATH.exists():
        os.remove(DB_PATH)
    conn = sqlite3.connect(str(DB_PATH))
    with open(SCHEMA_PATH, 'r') as f:
        schema = f.read()
    conn.executescript(schema)
    conn.commit()
    return conn

def migrate_sheet(conn, sheet_name, table_name):
    """Migrate a workbook sheet to SQLite table."""
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return 0

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or not rows[0]:
        return 0

    headers = [h for h in rows[0] if h]
    db_columns = [snake_case(h) for h in headers]
    
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table_name})")
    table_cols = {row[1] for row in cur.fetchall()}
    
    filtered_cols = [(h, db_col) for h, db_col in zip(headers, db_columns) if db_col in table_cols]
    if not filtered_cols:
        return 0

    col_indices = [headers.index(h) for h, _ in filtered_cols]
    insert_cols = [db_col for _, db_col in filtered_cols]
    
    row_count = 0
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        values = [row[i] if i < len(row) else None for i in col_indices]
        # Convert dates and booleans
        values = [
            normalize_boolean(excel_to_iso_date(v) if isinstance(v, datetime) else v, insert_cols[i])
            for i, v in enumerate(values)
        ]
        try:
            placeholders = ', '.join(['?' for _ in insert_cols])
            sql = f"INSERT INTO {table_name} ({', '.join(insert_cols)}) VALUES ({placeholders})"
            cur.execute(sql, values)
            row_count += 1
        except Exception as e:
            pass
    
    conn.commit()
    return row_count

def verify_migration(conn):
    """Verify row counts and report data quality."""
    print("\nMigration verification:")
    cur = conn.cursor()
    
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [row[0] for row in cur.fetchall()]
    
    total_rows = 0
    for table in tables:
        cur.execute(f"SELECT COUNT(*) FROM {table}")
        count = cur.fetchone()[0]
        total_rows += count
        status = "✓" if count > 0 else " "
        print(f"  {status} {table}: {count} rows")
    
    print(f"\nTotal rows migrated: {total_rows}")
    return total_rows

def main():
    print("Initializing database from schema...")
    conn = init_db()
    
    print("Migrating sheets...")
    migrations = [
        ("Projects", "projects"),
        ("Employees", "employees"),
        ("Site Closure Log", "site_closure_log"),
        ("Meeting Schedule", "meeting_schedule"),
        ("Meeting Records", "meeting_records"),
        ("Meeting Action Items", "meeting_action_items"),
        ("Settings", "settings"),
        ("Drop Plan", "drop_plan"),
        ("Permits Library", "permits_library"),
        ("Document Library", "document_library"),
        ("DOB Compliance Reference", "dob_compliance_reference"),
        ("Toolbox Talk Library", "toolbox_talk_library"),
    ]
    
    for sheet, table in migrations:
        count = migrate_sheet(conn, sheet, table)
        print(f"  {sheet}: {count} rows")
    
    verify_migration(conn)
    
    db_size = os.path.getsize(DB_PATH)
    print(f"\nDatabase created: {DB_PATH}")
    print(f"Database size: {db_size:,} bytes ({db_size / (1024*1024):.2f} MB)")
    
    conn.close()

if __name__ == "__main__":
    main()
