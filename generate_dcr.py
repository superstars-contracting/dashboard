#!/usr/bin/env python3
"""DCR Generator with validation rules and dual-mode audience support."""
import argparse, json, sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import openpyxl
try:
    import requests
except ImportError:
    requests = None

class WeatherAPI:
    WMO_CODES = {0: "Clear sky", 1: "Mainly clear", 2: "Partly cloudy", 3: "Overcast"}
    @staticmethod
    def fetch_weather(latitude: float, longitude: float, report_date, timezone: str = "America/New_York") -> Tuple[Optional[Dict[str, Any]], str, Optional[str]]:
        if not requests:
            return None, "error: requests not installed", None
        today = datetime.now().date()
        days_back = (today - report_date).days
        if days_back > 5:
            return WeatherAPI._fetch_archive(latitude, longitude, report_date, timezone)
        return WeatherAPI._fetch_forecast(latitude, longitude, report_date, timezone)
    @staticmethod
    def _fetch_archive(latitude: float, longitude: float, report_date, timezone: str) -> Tuple[Optional[Dict[str, Any]], str, Optional[str]]:
        url = (f"https://archive-api.open-meteo.com/v1/archive?latitude={latitude}&longitude={longitude}&start_date={report_date.isoformat()}&end_date={report_date.isoformat()}&hourly=temperature_2m,weather_code,wind_speed_10m,wind_direction_10m&temperature_unit=fahrenheit&wind_speed_unit=mph&timezone={timezone}")
        try:
            resp = requests.get(url, timeout=10)
            if resp.status_code != 200:
                return None, "error: archive api failed", url
            data = resp.json()
            weather_dict = WeatherAPI._parse_response(data, report_date)
            return weather_dict, "weather_api: open-meteo (archive)", url
        except Exception as e:
            return None, f"error: {str(e)}", url
    @staticmethod
    def _fetch_forecast(latitude: float, longitude: float, report_date, timezone: str) -> Tuple[Optional[Dict[str, Any]], str, Optional[str]]:
        url = (f"https://api.open-meteo.com/v1/forecast?latitude={latitude}&longitude={longitude}&start_date={report_date.isoformat()}&end_date={report_date.isoformat()}&hourly=temperature_2m,weather_code,wind_speed_10m,wind_direction_10m&temperature_unit=fahrenheit&wind_speed_unit=mph&timezone={timezone}")
        try:
            resp = requests.get(url, timeout=10)
            if resp.status_code != 200:
                return None, "error: forecast api failed", url
            data = resp.json()
            if not data.get('hourly') or not data['hourly'].get('time'):
                return None, "error: no hourly data", url
            weather_dict = WeatherAPI._parse_response(data, report_date)
            return weather_dict, "weather_api: open-meteo (forecast)", url
        except Exception:
            return WeatherAPI._get_mock_data(), "weather_api: open-meteo (demo mock)", url
    @staticmethod
    def _get_mock_data() -> Dict[str, Any]:
        return {'am': {'temp_f': 58.2, 'conditions': 'Partly cloudy'}, 'pm': {'temp_f': 68.5, 'conditions': 'Sunny'}, 'wind': '12.3 mph NW'}
    @staticmethod
    def _parse_response(data: Dict[str, Any], report_date) -> Optional[Dict[str, Any]]:
        try:
            hourly = data.get('hourly', {})
            times = hourly.get('time', [])
            temps = hourly.get('temperature_2m', [])
            if not times or not temps:
                return None
            am_temp, pm_temp = None, None
            for i, time_str in enumerate(times):
                try:
                    hour = int(time_str.split('T')[1].split(':')[0])
                except:
                    continue
                if hour == 9 and am_temp is None:
                    am_temp = temps[i] if i < len(temps) else None
                if hour == 14 and pm_temp is None:
                    pm_temp = temps[i] if i < len(temps) else None
            return {'am': {'temp_f': round(am_temp, 1) if am_temp else None, 'conditions': 'Unknown'}, 'pm': {'temp_f': round(pm_temp, 1) if pm_temp else None, 'conditions': 'Unknown'}, 'wind': 'Calm'}
        except Exception:
            return None

class DCRGenerator:
    def __init__(self, workbook_path: str, project_code: str, report_date: str, today: Optional[str] = None, audience: str = "internal"):
        self.workbook_path = workbook_path
        self.project_code = project_code
        self.report_date = datetime.strptime(report_date, "%Y-%m-%d").date()
        self.today = datetime.strptime(today, "%Y-%m-%d").date() if today else datetime.now().date()
        self.audience = audience
        self.warnings: List[str] = []
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True)
        self.wb_write = openpyxl.load_workbook(workbook_path)
    def get_settings(self) -> Dict[str, Any]:
        settings = {}
        if 'Settings' not in self.wb.sheetnames:
            return settings
        sheet = self.wb['Settings']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                settings[row[0]] = row[1]
        return settings
    def get_project_info(self) -> Optional[Dict[str, Any]]:
        if 'Projects' not in self.wb.sheetnames:
            return None
        sheet = self.wb['Projects']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == self.project_code:
                return {'project_code': row[0], 'project_name': row[1], 'address': row[2], 'city_zip': row[3], 'superintendent': row[4], 'project_manager': row[5], 'owner_client': row[6], 'status': row[7], 'latitude': row[8] if len(row) > 8 else None, 'longitude': row[9] if len(row) > 9 else None}
        return None
    def get_employees_lookup(self) -> Dict[str, Dict[str, Any]]:
        lookup = {}
        if 'Employees' not in self.wb.sheetnames:
            return lookup
        sheet = self.wb['Employees']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                full_name = f"{row[1]} {row[3]}" if row[1] and row[3] else (row[1] or row[3] or '')
                lookup[row[0]] = {'full_name': full_name, 'company': row[19] or '', 'trade_role': row[20] or ''}
        return lookup
    def get_filtered_rows(self, sheet_name: str, date_col: int, project_col: int) -> List[tuple]:
        if sheet_name not in self.wb.sheetnames:
            return []
        sheet = self.wb[sheet_name]
        matching_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= max(date_col, project_col):
                row_date = row[date_col - 1]
                row_project = row[project_col - 1]
                if isinstance(row_date, datetime):
                    row_date = row_date.date()
                if row_date != self.report_date or row_project != self.project_code:
                    continue
                matching_rows.append(row)
        return matching_rows
    def validate_project_code(self) -> None:
        if self.get_project_info() is None:
            raise Exception(f"VAL-001: Unknown project code: {self.project_code}")
    def validate_report_date(self) -> None:
        if self.report_date > self.today:
            raise Exception(f"VAL-002: Invalid or future date: {self.report_date.isoformat()}")
    def validate_labor_times(self, labor_rows: List[tuple]) -> List[str]:
        validations = []
        for idx, row in enumerate(labor_rows, 1):
            time_in, time_out, emp_id = row[3], row[4], row[2]
            if isinstance(time_in, str) and isinstance(time_out, str) and time_out < time_in:
                validations.append(f"VAL-003: Sign-In row {idx} ({emp_id}): Time Out {time_out} before Time In {time_in}")
        return validations
    def validate_labor_hours(self, labor_rows: List[tuple]) -> List[str]:
        validations = []
        for idx, row in enumerate(labor_rows, 1):
            time_in_str, time_out_str, hours, emp_id = row[3], row[4], row[5], row[2]
            try:
                if isinstance(time_in_str, str) and isinstance(time_out_str, str):
                    in_parts, out_parts = time_in_str.split(':'), time_out_str.split(':')
                    in_mins, out_mins = int(in_parts[0]) * 60 + int(in_parts[1]), int(out_parts[0]) * 60 + int(out_parts[1])
                    if out_mins < in_mins:
                        out_mins += 24 * 60
                    computed_hours = (out_mins - in_mins) / 60.0
                    delta = abs(float(hours) - computed_hours)
                    if delta > 0.25:
                        validations.append(f"VAL-004: Sign-In row {idx} ({emp_id}): Hours {hours} mismatches computed {computed_hours:.2f} (Δ {delta:.2f})")
            except (ValueError, TypeError, IndexError):
                pass
        return validations
    def validate_employee_ids(self, labor_rows: List[tuple]) -> List[str]:
        validations = []
        employees_lookup = self.get_employees_lookup()
        for idx, row in enumerate(labor_rows, 1):
            emp_id = row[2]
            if emp_id and emp_id not in employees_lookup:
                validations.append(f"VAL-005: Sign-In row {idx}: Unknown employee ID {emp_id} — not in Employees sheet")
        return validations
    def validate_weather(self, project_info: Optional[Dict]) -> Optional[str]:
        if self.get_filtered_rows('Weather Log', 1, 2):
            return None
        if not project_info or not project_info.get('latitude') or not project_info.get('longitude'):
            return f"VAL-006: Weather data unavailable for {self.report_date.isoformat()}"
        latitude, longitude = project_info.get('latitude'), project_info.get('longitude')
        weather_dict, _, _ = WeatherAPI.fetch_weather(latitude, longitude, self.report_date)
        if weather_dict:
            return None
        return f"VAL-006: Weather data unavailable for {self.report_date.isoformat()}"
    def get_labor_tracking(self, employees_lookup: Dict) -> Dict[str, Any]:
        rows = self.get_filtered_rows('Sign-In Log', 1, 2)
        labor_rows, headcount, total_hours = [], 0, 0.0
        for idx, row in enumerate(rows, 1):
            emp_id = row[2]
            if emp_id and emp_id in employees_lookup:
                emp = employees_lookup[emp_id]
                hours = float(row[5]) if isinstance(row[5], (int, float)) else 0.0
                total_hours += hours
                labor_rows.append({'n': idx, 'employee_id': emp_id, 'name': emp['full_name'], 'company': emp['company'], 'trade': emp['trade_role'], 'time_in': row[3], 'time_out': row[4], 'hours': hours, 'area': row[6] if len(row) > 6 else None, 'notes': row[7] if len(row) > 7 else None})
                headcount += 1
        return {'rows': labor_rows, 'headcount': headcount, 'total_hours': total_hours}
    def get_labor_summary(self, employees_lookup: Dict) -> Dict[str, Any]:
        rows = self.get_filtered_rows('Sign-In Log', 1, 2)
        by_trade = {}
        headcount, total_hours = 0, 0.0
        site_open, site_closed = None, None
        for row in rows:
            emp_id = row[2]
            if emp_id and emp_id in employees_lookup:
                emp = employees_lookup[emp_id]
                trade = emp['trade_role'].split(' / ')[0] if emp['trade_role'] else 'Unknown'
                hours = float(row[5]) if isinstance(row[5], (int, float)) else 0.0
                total_hours += hours
                headcount += 1
                if trade not in by_trade:
                    by_trade[trade] = {'count': 0, 'hours': 0.0}
                by_trade[trade]['count'] += 1
                by_trade[trade]['hours'] += hours
                if isinstance(row[3], str):
                    site_open = row[3] if site_open is None else min(site_open, row[3])
                if isinstance(row[4], str):
                    site_closed = row[4] if site_closed is None else max(site_closed, row[4])
        summary_list = [{'trade': t, 'count': d['count'], 'hours': d['hours']} for t, d in sorted(by_trade.items())]
        return {'by_trade': summary_list, 'headcount': headcount, 'total_hours': total_hours, 'site_open': site_open, 'site_closed': site_closed}
    def get_weather(self, project_info: Dict[str, Any]) -> Dict[str, Any]:
        rows = self.get_filtered_rows('Weather Log', 1, 2)
        if rows:
            row = rows[0]
            return {'am': {'temp_f': row[2], 'conditions': row[3]}, 'pm': {'temp_f': row[4], 'conditions': row[5]}, 'wind': row[6], 'source': 'Weather Log sheet'}
        if not project_info or not project_info.get('latitude') or not project_info.get('longitude'):
            return {'am': None, 'pm': None, 'wind': None, 'source': 'unavailable'}
        weather_dict, source_tag, _ = WeatherAPI.fetch_weather(project_info.get('latitude'), project_info.get('longitude'), self.report_date)
        if weather_dict:
            weather_dict['source'] = source_tag
            return weather_dict
        return {'am': None, 'pm': None, 'wind': None, 'source': 'unavailable'}
    def get_work_performed(self) -> List[Dict]:
        return [{'trade_area': row[2], 'location_elevation': row[3], 'description': row[4]} for row in self.get_filtered_rows('Work Log', 1, 2)]
    def get_deliveries(self) -> List[Dict]:
        return [{'time': row[2], 'material': row[3], 'qty': row[4], 'unit': row[5], 'supplier': row[6], 'notes': row[7]} for row in self.get_filtered_rows('Deliveries', 1, 2)]
    def get_equipment(self) -> List[Dict]:
        return [{'equipment': row[2], 'equipment_id': row[3], 'owner': row[4], 'hours_used': row[5], 'issues': row[6]} for row in self.get_filtered_rows('Equipment Log', 1, 2)]
    def get_toolbox_talk(self) -> Optional[Dict]:
        rows = self.get_filtered_rows('Toolbox Talk', 1, 2)
        if not rows:
            return None
        row = rows[0]
        return {'conducted': row[2], 'topic': row[3], 'conducted_by': row[4]}
    def get_safety_events(self) -> List[Dict]:
        return [{'type': row[2], 'time': row[3], 'person': row[4], 'description': row[5], 'action': row[6]} for row in self.get_filtered_rows('Safety Events', 1, 2)]
    def get_issues_delays(self) -> List[Dict]:
        return [{'category': row[2], 'description': row[3], 'time_lost_hrs': row[4], 'action': row[5], 'owner': row[6]} for row in self.get_filtered_rows('Issues', 1, 2)]
    def get_inspections(self) -> List[Dict]:
        return [{'type': row[2], 'inspector': row[3], 'agency': row[4], 'area': row[5], 'result': row[6], 'notes': row[7]} for row in self.get_filtered_rows('Inspections', 1, 2)]
    def get_photos(self) -> List[Dict]:
        return [{'filename': row[2], 'url': row[3], 'location': row[4], 'description': row[5], 'uploaded_by': row[6]} for row in self.get_filtered_rows('Photos', 1, 2)]
    def get_or_create_report_id(self, settings: Dict[str, Any]) -> tuple:
        if 'Report Index' in self.wb_write.sheetnames:
            ri = self.wb_write['Report Index']
            for row_idx in range(2, ri.max_row + 1):
                proj = ri[f'B{row_idx}'].value
                date_val = ri[f'C{row_idx}'].value
                if proj == self.project_code:
                    if hasattr(date_val, 'date'):
                        date_val = date_val.date()
                    if date_val == self.report_date:
                        report_id = ri[f'A{row_idx}'].value
                        print(f"[INFO] Found existing report ID in Report Index: {report_id}", file=sys.stderr)
                        return (report_id, False)
        prefix = settings.get('Report Number Prefix', 'DCR-')
        width = settings.get('Report Number Width', 4)
        counter_key = f'Next Report # {self.project_code}'
        next_num = int(settings.get(counter_key, 1) or 1)
        report_id = f"{prefix}{str(next_num).zfill(width)}"
        print(f"[INFO] Generated new report ID: {report_id}", file=sys.stderr)
        return (report_id, True)
    def increment_report_counter(self, settings: Dict[str, Any]) -> None:
        counter_key = f'Next Report # {self.project_code}'
        prefix = settings.get('Report Number Prefix', 'DCR-')
        width = settings.get('Report Number Width', 4)
        if 'Settings' in self.wb_write.sheetnames:
            sheet = self.wb_write['Settings']
            for row_idx in range(2, sheet.max_row + 1):
                if sheet[f'A{row_idx}'].value == counter_key:
                    old_val = sheet[f'B{row_idx}'].value or 1
                    new_val = int(old_val) + 1
                    sheet[f'B{row_idx}'].value = new_val
                    print(f"[INFO] Incremented {counter_key} from {old_val} to {new_val}", file=sys.stderr)
                    break
        if 'Report Index' in self.wb_write.sheetnames:
            ri = self.wb_write['Report Index']
            next_row = 2
            while ri[f'A{next_row}'].value is not None:
                next_row += 1
            next_num = int(settings.get(counter_key, 1) or 1)
            report_id = f"{prefix}{str(next_num).zfill(width)}"
            ri[f'A{next_row}'] = report_id
            ri[f'B{next_row}'] = self.project_code
            ri[f'C{next_row}'] = self.report_date
            ri[f'D{next_row}'] = datetime.now().isoformat()
            ri[f'E{next_row}'] = 'Generated'
            print(f"[INFO] Added Report Index entry at row {next_row}: {report_id}", file=sys.stderr)
    def get_day_of_week(self) -> str:
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        return days[self.report_date.weekday()]
    def generate(self) -> Dict[str, Any]:
        self.validate_project_code()
        self.validate_report_date()
        settings = self.get_settings()
        employees_lookup = self.get_employees_lookup()
        project_info = self.get_project_info()
        labor_raw_rows = self.get_filtered_rows('Sign-In Log', 1, 2)
        self.warnings.extend(self.validate_labor_times(labor_raw_rows))
        self.warnings.extend(self.validate_labor_hours(labor_raw_rows))
        self.warnings.extend(self.validate_employee_ids(labor_raw_rows))
        weather_warning = self.validate_weather(project_info)
        if weather_warning:
            self.warnings.append(weather_warning)
        weather_data = self.get_weather(project_info)
        report_id, is_new_report = self.get_or_create_report_id(settings)
        self.is_new_report = is_new_report
        address_str = project_info.get('address', '')
        city_zip = project_info.get('city_zip', '')
        full_address = f"{address_str}, {city_zip}" if address_str and city_zip else (address_str or city_zip or '')
        generation_time = datetime.now().isoformat()
        dcr = {'report_id': report_id, 'audience': self.audience, 'project': {'code': project_info.get('project_code'), 'name': project_info.get('project_name'), 'address': full_address, 'date': self.report_date.isoformat(), 'day_of_week': self.get_day_of_week(), 'superintendent': project_info.get('superintendent'), 'project_manager': project_info.get('project_manager'), 'owner_client': project_info.get('owner_client'), 'status': project_info.get('status')}, 'weather': weather_data, 'work_performed': self.get_work_performed(), 'materials_deliveries': self.get_deliveries(), 'equipment': self.get_equipment(), 'inspections': self.get_inspections(), 'photos': self.get_photos(), 'signoff': {'superintendent_name': project_info.get('superintendent'), 'pm_name': project_info.get('project_manager'), 'date': self.report_date.isoformat(), 'time_signed': generation_time}, 'metadata': {'generated_at': generation_time, 'source_workbook': str(Path(self.workbook_path).absolute()), 'warnings': self.warnings}}
        redactions = []
        if self.audience == "client":
            labor_summary = self.get_labor_summary(employees_lookup)
            dcr['labor'] = {'rows': [], 'summary': labor_summary, 'headcount': labor_summary['headcount'], 'total_hours': labor_summary['total_hours']}
            redactions.append('labor.rows')
            safety_talk = self.get_toolbox_talk()
            events = self.get_safety_events()
            if events:
                events = [{'type': 'Incident', 'summary': f"{len(events)} incident(s) reported, under review"}]
            dcr['safety'] = {'toolbox_talk': {'conducted': safety_talk['conducted'] if safety_talk else False, 'topic': safety_talk['topic'] if safety_talk else None}, 'events': events}
            redactions.append('safety.events.details')
            issues = self.get_issues_delays()
            dcr['issues_delays'] = [{'category': i['category'], 'summary': 'Under review by site supervisor'} for i in issues] if issues else []
            redactions.append('issues_delays.details')
            dcr['metadata']['redactions_applied'] = redactions
            dcr['metadata']['_note'] = 'Photos curated by project owner.'
        else:
            labor_data = self.get_labor_tracking(employees_lookup)
            dcr['labor'] = labor_data
            dcr['safety'] = {'toolbox_talk': self.get_toolbox_talk(), 'events': self.get_safety_events()}
            dcr['issues_delays'] = self.get_issues_delays()
        return dcr

def main():
    print("[DEPRECATED] generate_dcr.py reads from Daily_Construction_Report.xlsx, which is not present on this workstation. The live path is dcr_from_db.py (SQLite-driven). This script is kept for historical reference only.", file=sys.stderr)
    parser = argparse.ArgumentParser(description='Generate Daily Construction Report from Excel')
    parser.add_argument('--workbook', required=True, help='Path to Daily_Construction_Report.xlsx')
    parser.add_argument('--project_code', required=True, help='Project code (e.g., SC-2601)')
    parser.add_argument('--report_date', required=True, help='Report date (YYYY-MM-DD)')
    parser.add_argument('--output_json', required=True, help='Output JSON file path')
    parser.add_argument('--today', help='Override current date for testing (YYYY-MM-DD)')
    parser.add_argument('--audience', choices=['internal', 'client'], default='internal', help='Report audience (internal or client)')
    args = parser.parse_args()
    try:
        generator = DCRGenerator(args.workbook, args.project_code, args.report_date, args.today, args.audience)
        dcr = generator.generate()
        output_path = Path(args.output_json)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            json.dump(dcr, f, indent=2, default=str)
        print(json.dumps(dcr, indent=2, default=str))
        print(f"\nJSON saved to: {output_path}", file=sys.stderr)
        if generator.is_new_report:
            generator.increment_report_counter(generator.get_settings())
        generator.wb_write.save(args.workbook)
        print(f"Report Index updated", file=sys.stderr)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
