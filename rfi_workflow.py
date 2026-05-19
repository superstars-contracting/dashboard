#!/usr/bin/env python3
"""
RFI Workflow Orchestrator
Reads RFI Log from workbook, applies routing rules, computes SLA deadlines,
transitions status, drafts emails, generates PDFs, and updates audit logs.
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta, date
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import subprocess

# Recipient email mapping
RECIPIENT_MAP = {
    'Architect': 'arch@nyparkdesigns.com',
    'Structural Engineer': 'hartwick@hartwickeng.com',
    'MEP Engineer': 'mep@nyparkdesigns.com',
    'Project Manager': 'susan.park@superstars.nyc',
    'Superintendent': 'robert.caldwell@superstars.nyc',
    'Owner Rep': 'jhellman@acmerealty.com',
    'Engineer': 'hartwick@hartwickeng.com',
}

RECIPIENT_NAMES = {
    'Architect': 'Architect',
    'Structural Engineer': 'Structural Engineer',
    'MEP Engineer': 'MEP Engineer',
    'Project Manager': 'Susan Park',
    'Superintendent': 'Robert Caldwell',
    'Owner Rep': 'J. Hellman',
    'Engineer': 'Structural Engineer',
}

# SLA definitions (hours to days)
SLA_MAP = {
    'Urgent': 24,
    'High': 48,
    'Medium': 120,
    'Low': 168,
}

# Routing rules: (predicate, recipients, notification_type)
ROUTING_RULES = [
    (lambda r: r['priority'] == 'Urgent',
     ['Architect', 'Engineer', 'Project Manager', 'Owner Rep', 'Superintendent'],
     'high_priority'),
    (lambda r: r['category'] in ['Architectural', 'Facade'],
     ['Architect', 'Project Manager'],
     'normal'),
    (lambda r: r['category'] == 'Structural',
     ['Structural Engineer', 'Project Manager'],
     'normal'),
    (lambda r: r['category'] == 'MEP',
     ['MEP Engineer', 'Project Manager'],
     'normal'),
    (lambda r: r['category'] in ['Site Condition', 'Other'],
     ['Project Manager', 'Superintendent'],
     'normal'),
]


class RFIWorkflow:
    def __init__(self, workbook_path, today_str, actions_dir):
        self.workbook_path = workbook_path
        self.today = datetime.fromisoformat(today_str).date()
        self.actions_dir = Path(actions_dir)
        self.actions_dir.mkdir(parents=True, exist_ok=True)

        self.run_id = f"RUN-{self.today}-001"
        self.started_at = datetime.now()
        self.actions = []
        self.summary = {
            'rfis_routed': 0,
            'emails_drafted': 0,
            'pdfs_generated': 0,
            'escalations': 0,
            'overdue_count': 0,
            'urgent_count': 0,
        }

        # Create subdirectories
        (self.actions_dir / 'email_drafts').mkdir(exist_ok=True)
        (self.actions_dir / 'pdf').mkdir(exist_ok=True)

    def load_rfis(self):
        """Load all RFIs from RFI Log sheet."""
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb['RFI Log']

        # Map column headers
        headers = {cell.value: idx for idx, cell in enumerate(ws[1])}

        rfis = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] is None:
                continue

            rfi = {
                'rfi_id': row[headers['RFI Number']],
                'project_code': row[headers['Project Code']],
                'date_submitted': str(row[headers['Date Submitted']]) if row[headers['Date Submitted']] else None,
                'title': row[headers['Title']],
                'category': row[headers['Category']],
                'priority': row[headers['Priority']],
                'status': row[headers['Status']],
                'due_date': str(row[headers['Response Due Date']]) if row[headers['Response Due Date']] else None,
            }
            rfis.append(rfi)

        wb.close()
        return rfis

    def apply_routing_rules(self, rfi):
        """Determine recipient list based on routing rules."""
        recipients = set()
        notification_type = 'normal'

        # First pass: apply category-based rules
        for predicate, rule_recipients, rule_notif in ROUTING_RULES[1:]:  # Skip Urgent rule
            try:
                if predicate(rfi):
                    recipients.update(rule_recipients)
                    notification_type = rule_notif
                    break  # Stop at first matching category rule
            except:
                continue

        # Second pass: Urgent override ADDS to recipients + escalates notification
        if rfi['priority'] == 'Urgent':
            recipients.update(['Architect', 'Structural Engineer', 'Project Manager', 'Owner Rep', 'Superintendent'])
            notification_type = 'high_priority'

        return sorted(list(recipients)), notification_type

    def compute_sla(self, rfi):
        """Compute response due date from submission + SLA hours."""
        priority = rfi.get('priority', 'Medium')
        sla_hours = SLA_MAP.get(priority, 120)

        try:
            date_submitted = datetime.fromisoformat(rfi['date_submitted']).date()
        except:
            return None, False, None

        due_date = date_submitted + timedelta(hours=sla_hours)
        past_due = self.today > due_date
        past_due_hours = None

        if past_due:
            past_due_hours = int((self.today - due_date).total_seconds() / 3600)

        return str(due_date), past_due, past_due_hours

    def transition_status(self, rfi, due_date_str, past_due):
        """Apply status transition rules."""
        status = rfi['status']

        if status == 'Draft':
            return status

        if past_due and status in ['Submitted', 'Under Review']:
            return 'Overdue'

        if status == 'Under Review' and past_due:
            return 'Overdue'

        return status

    def draft_email(self, rfi, recipient):
        """Generate branded HTML email notification for a recipient."""
        priority = rfi.get('priority', 'Medium')
        is_urgent = priority == 'Urgent'

        subject = f"{'URGENT RFI - ' if is_urgent else 'New RFI Submitted - '}{rfi['rfi_id']} - {rfi['title'][:40]}"

        # Compute due date countdown
        try:
            due_date = datetime.fromisoformat(rfi['due_date']).date() if rfi.get('due_date') else None
            if due_date:
                days_until = (due_date - self.today).days
                if days_until < 0:
                    countdown_str = f"<span style='color:#B11E2E; font-weight:700;'>OVERDUE by {abs(days_until)} days</span>"
                elif days_until == 0:
                    countdown_str = f"<span style='color:#B11E2E;'>Due TODAY</span>"
                else:
                    countdown_str = f"Response due {due_date.strftime('%b %d')} · {days_until} days from now"
            else:
                countdown_str = "No due date set"
        except:
            countdown_str = "No due date set"

        recipient_name = RECIPIENT_NAMES.get(recipient, recipient)

        # Priority badge color mapping
        badge_colors = {
            'Urgent': ('background-color:#B11E2E; color:#fff;', '#B11E2E'),
            'High': ('background-color:#d75a5a; color:#fff;', '#d75a5a'),
            'Medium': ('background-color:#d4af37; color:#1a1a1a;', '#d4af37'),
            'Low': ('background-color:#888; color:#fff;', '#888'),
        }
        badge_style, _ = badge_colors.get(priority, badge_colors['Medium'])

        # Build star SVG inline
        star_svg = '''<svg width="16" height="16" viewBox="0 0 24 24" fill="#B11E2E" style="display:inline; margin-right:4px; vertical-align:middle;">
  <polygon points="12,2 15.09,10.26 23.36,10.26 17.54,15.64 19.63,23.91 12,18.53 4.37,23.91 6.46,15.64 0.64,10.26 8.91,10.26"/>
</svg>'''

        # Urgent callout (shown above card for urgent RFIs)
        urgent_callout = ""
        if is_urgent:
            urgent_callout = f'''<tr><td style="padding: 0 20px 20px;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background-color:#B11E2E; border-radius:4px; padding:16px;">
    <tr><td style="color:#fff; font-size:15px; font-weight:700;">
      URGENT — Work Stoppage Risk. Please respond as soon as possible.
    </td></tr>
  </table>
</td></tr>'''

        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{subject}</title>
</head>
<body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; background-color:#f5f5f5; margin:0; padding:20px;">

<table width="100%" cellpadding="0" cellspacing="0" style="max-width:600px; margin:0 auto; background-color:#fff;">

  <!-- Letterhead -->
  <tr><td style="background-color:#14161C; padding:24px 20px; text-align:center; border-bottom:2px solid #B11E2E;">
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr><td style="text-align:center;">
        {star_svg}
        <span style="color:#fff; font-size:16px; font-weight:700; letter-spacing:1px;">SUPERSTARS CONTRACTING</span>
      </td></tr>
      <tr><td style="text-align:center; color:#999; font-size:11px; margin-top:4px; padding-top:8px;">Project Notification</td></tr>
    </table>
  </td></tr>

  <!-- Greeting + intro -->
  <tr><td style="padding:24px 20px 16px;">
    <p style="margin:0 0 12px; font-size:15px; color:#1a1a1a;">Hi {recipient_name},</p>
    <p style="margin:0; font-size:14px; color:#555; line-height:1.5;">A new RFI has been submitted for review on the 890 E 135th St 890 E 135th Street project. Brief details below — full document attached as PDF.</p>
  </td></tr>

  {urgent_callout}

  <!-- Info card -->
  <tr><td style="padding:0 20px 20px;">
    <table width="100%" cellpadding="16" cellspacing="0" style="border:1px solid #ddd; border-left:4px solid #B11E2E; background-color:#fafafa; border-radius:4px;">
      <tr><td>
        <div style="font-family: 'Playfair Display', Georgia, serif; font-size:16px; font-weight:600; color:#1a1a1a; margin-bottom:8px;">
          {rfi['rfi_id']} — {rfi['title']}
        </div>
        <div style="margin-bottom:12px;">
          <span style="{badge_style} padding:4px 10px; border-radius:3px; font-size:11px; font-weight:700;">{priority}</span>
        </div>
        <div style="font-size:12px; color:#666; line-height:1.6; margin-bottom:12px;">
          <div>Submitted by: <span style="color:#1a1a1a; font-weight:600;">{rfi.get('category', 'N/A')}</span></div>
          <div style="margin-top:8px;">Response due: {countdown_str}</div>
        </div>
      </td></tr>
    </table>
  </td></tr>

  <!-- Action prompt -->
  <tr><td style="padding:0 20px 24px; font-size:13px; color:#555; line-height:1.5;">
    Please review the attached PDF and respond by replying to this email or via the project portal.
  </td></tr>

  <!-- Footer -->
  <tr><td style="background-color:#f5f5f5; padding:16px 20px; text-align:center; border-top:1px solid #ddd; font-size:10px; color:#999;">
    {star_svg}
    <span style="font-size:10px; color:#999;">Generated by Superstars Contracting Project Console</span>
  </td></tr>

</table>

</body>
</html>"""

        return html

    def generate_pdf(self, rfi_id):
        """Generate PDF from RFI HTML using weasyprint or graceful fallback."""
        html_file = self.actions_dir.parent / f"{rfi_id}.html"
        pdf_file = self.actions_dir / 'pdf' / f"{rfi_id}.pdf"

        if not html_file.exists():
            return False, "HTML file not found"

        try:
            from weasyprint import HTML
            HTML(str(html_file)).write_pdf(str(pdf_file))
            return True, str(pdf_file)
        except ImportError:
            return False, "weasyprint unavailable"
        except Exception as e:
            return False, f"weasyprint error: {str(e)}"

    def run(self):
        """Execute the complete workflow."""
        # Stage 1: Load RFIs
        rfis = self.load_rfis()
        self.log_action('workflow', 'loaded', f"Loaded {len(rfis)} RFIs from workbook")

        # Stage 2-7: Process each RFI
        for rfi in rfis:
            self.process_rfi(rfi)

        # Write workbook updates
        self.update_workbook(rfis)

        # Write audit log
        self.write_audit_log()

        # Write dashboard state
        self.write_dashboard_state(rfis)

    def process_rfi(self, rfi):
        """Process a single RFI through all stages."""
        rfi_id = rfi['rfi_id']

        # Stage 1: Load
        self.log_action(rfi_id, 'loaded', {
            'status_before': rfi['status'],
            'priority': rfi['priority'],
            'category': rfi['category'],
        })

        # Stage 2: Apply routing rules
        recipients, notif_type = self.apply_routing_rules(rfi)
        self.log_action(rfi_id, 'routing_applied', {
            'recipients': recipients,
            'notification_type': notif_type,
        })

        # Stage 3: Compute SLA
        due_date_str, past_due, past_due_hours = self.compute_sla(rfi)
        if due_date_str:
            rfi['due_date'] = due_date_str
            self.log_action(rfi_id, 'sla_computed', {
                'priority': rfi['priority'],
                'sla_hours': SLA_MAP.get(rfi['priority'], 0),
                'due_date': due_date_str,
                'past_due': past_due,
                'past_due_hours': past_due_hours,
            })

        # Stage 4: State transition
        old_status = rfi['status']
        new_status = self.transition_status(rfi, due_date_str, past_due)
        rfi['status'] = new_status

        if new_status != old_status:
            self.log_action(rfi_id, 'status_transitioned', {
                'status_before': old_status,
                'status_after': new_status,
            })

        if new_status == 'Overdue':
            self.summary['overdue_count'] += 1
            self.log_action(rfi_id, 'escalated_to_pm', {
                'reason': f"Past due {past_due_hours // 24} days",
            })
            self.summary['escalations'] += 1

        if rfi['priority'] == 'Urgent':
            self.summary['urgent_count'] += 1

        # Stage 5a: Generate PDFs
        pdf_ok, pdf_msg = self.generate_pdf(rfi_id)
        if pdf_ok:
            self.log_action(rfi_id, 'pdf_generated', {'file': pdf_msg})
            self.summary['pdfs_generated'] += 1
        else:
            self.log_action(rfi_id, 'pdf_generation_failed', {'reason': pdf_msg})

        # Stage 5b: Draft emails
        for recipient in recipients:
            email_html = self.draft_email(rfi, recipient)
            recipient_slug = recipient.replace(' ', '_')
            email_file = self.actions_dir / 'email_drafts' / f"{rfi_id}-to-{recipient_slug}.html"

            with open(email_file, 'w') as f:
                f.write(email_html)

            self.log_action(rfi_id, 'email_drafted', {
                'recipient': recipient,
                'email': RECIPIENT_MAP.get(recipient, 'N/A'),
                'file': str(email_file.relative_to(self.actions_dir.parent)),
            })
            self.summary['emails_drafted'] += 1

        self.summary['rfis_routed'] += 1

    def log_action(self, entity, action, details=None):
        """Log a workflow action."""
        record = {
            'timestamp': datetime.now().isoformat(),
            'entity': entity,
            'action': action,
        }
        if details:
            record.update(details if isinstance(details, dict) else {'detail': str(details)})
        self.actions.append(record)

    def update_workbook(self, rfis):
        """Write status and due date changes back to workbook."""
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb['RFI Log']

        headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
        status_col = headers.get('Status')
        due_col = headers.get('Response Due Date')

        # Check if "Last Workflow Run" column exists
        last_run_col = None
        if 'Last Workflow Run' in headers:
            last_run_col = headers['Last Workflow Run']
        else:
            # Add new column
            last_run_col = ws.max_column + 1
            ws.cell(row=1, column=last_run_col, value='Last Workflow Run')

        for row_idx, rfi in enumerate(rfis, start=2):
            rfi_id = rfi['rfi_id']

            # Update status
            if status_col is not None:
                ws.cell(row=row_idx, column=status_col + 1, value=rfi['status'])

            # Update due date
            if due_col is not None and rfi.get('due_date'):
                ws.cell(row=row_idx, column=due_col + 1, value=rfi['due_date'])

            # Update last run timestamp
            ws.cell(row=row_idx, column=last_run_col + 1, value=datetime.now().isoformat())

        wb.save(self.workbook_path)
        wb.close()
        self.log_action('workflow', 'workbook_updated', f"Updated {len(rfis)} RFI rows")

    def write_audit_log(self):
        """Write audit log JSON."""
        log_data = {
            'run_id': self.run_id,
            'started_at': self.started_at.isoformat(),
            'completed_at': datetime.now().isoformat(),
            'today': str(self.today),
            'rfis_processed': self.summary['rfis_routed'],
            'actions': self.actions,
            'summary': self.summary,
        }

        log_file = self.actions_dir / 'rfi_workflow_log.json'
        with open(log_file, 'w') as f:
            json.dump(log_data, f, indent=2)

    def write_dashboard_state(self, rfis):
        """Write dashboard state JSON."""
        by_category = {}
        by_priority = {}
        by_status = {}
        rfi_list = []

        for rfi in rfis:
            cat = rfi['category']
            by_category[cat] = by_category.get(cat, 0) + 1

            pri = rfi['priority']
            by_priority[pri] = by_priority.get(pri, 0) + 1

            sta = rfi['status']
            by_status[sta] = by_status.get(sta, 0) + 1

            rfi_list.append({
                'rfi_id': rfi['rfi_id'],
                'title': rfi['title'],
                'status': rfi['status'],
                'priority': rfi['priority'],
                'category': rfi['category'],
            })

        state = {
            'last_updated': datetime.now().isoformat(),
            'metrics': {
                'open_rfis': len([r for r in rfis if r['status'] not in ['Closed', 'Response Received']]),
                'overdue_rfis': self.summary['overdue_count'],
                'urgent_rfis': self.summary['urgent_count'],
                'average_response_time_days': None,
                'by_category': by_category,
                'by_priority': by_priority,
                'by_status': by_status,
            },
            'rfis': rfi_list,
        }

        state_file = self.actions_dir / 'rfi_dashboard_state.json'
        with open(state_file, 'w') as f:
            json.dump(state, f, indent=2)


def main():
    parser = argparse.ArgumentParser(description='RFI Workflow Orchestrator')
    parser.add_argument('--workbook', required=True, help='Path to workbook')
    parser.add_argument('--today', required=True, help='Date (YYYY-MM-DD)')
    parser.add_argument('--actions-dir', required=True, help='Output directory for artifacts')

    args = parser.parse_args()

    workflow = RFIWorkflow(args.workbook, args.today, args.actions_dir)
    workflow.run()

    print(f"✓ Workflow complete")
    print(f"  RFIs processed: {workflow.summary['rfis_routed']}")
    print(f"  Emails drafted: {workflow.summary['emails_drafted']}")
    print(f"  PDFs generated: {workflow.summary['pdfs_generated']}")
    print(f"  Escalations: {workflow.summary['escalations']}")
    print(f"  Output: {args.actions_dir}")


if __name__ == '__main__':
    main()
