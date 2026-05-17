import json
import argparse
from openpyxl import load_workbook
from datetime import datetime

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--talk_id', required=True)
    parser.add_argument('--project_code', required=True)
    parser.add_argument('--date', required=True)
    parser.add_argument('--foreman_id', default=None)
    parser.add_argument('--output_json', required=True)
    args = parser.parse_args()

    wb = load_workbook(args.workbook)
    
    lib_sheet = wb['Toolbox Talk Library']
    projects_sheet = wb['Projects']
    employees_sheet = wb['Employees']
    
    talk_data = None
    for row in lib_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == args.talk_id:
            talk_data = row
            break
    
    if not talk_data:
        print(f"Talk {args.talk_id} not found")
        return
    
    project_data = None
    for row in projects_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == args.project_code:
            project_data = row
            break
    
    if not project_data:
        print(f"Project {args.project_code} not found")
        return
    
    foreman_name = None
    foreman_id = args.foreman_id or (project_data[4] if len(project_data) > 4 else None)
    if foreman_id:
        for row in employees_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == foreman_id:
                foreman_name = row[1]
                break
    
    output = {
        'talk_id': talk_data[0],
        'title': talk_data[1],
        'category': talk_data[2],
        'dob_reference': talk_data[3],
        'osha_reference': talk_data[4],
        'duration_estimated': talk_data[5],
        'required_for': talk_data[6],
        'frequency_recommendation': talk_data[7],
        'hazards_summary': talk_data[8],
        'key_practices': [p.strip() for p in str(talk_data[9]).split(';') if p.strip()],
        'required_ppe': [p.strip() for p in str(talk_data[10]).split(';') if p.strip()],
        'discussion_questions': [q.strip() for q in str(talk_data[11]).split(';') if q.strip()],
        'required_inspections': talk_data[12],
        'related_certifications': [c.strip() for c in str(talk_data[13]).split(';') if c.strip()],
        'project': {
            'code': project_data[0],
            'name': project_data[1],
            'address': project_data[2],
            'city_zip': project_data[3],
            'superintendent': project_data[4] if len(project_data) > 4 else None
        },
        'scheduled_for': args.date,
        'conducted_by_planned': foreman_name,
        'sign_in_required': True
    }
    
    with open(args.output_json, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"Generated {args.output_json}")

if __name__ == '__main__':
    main()
