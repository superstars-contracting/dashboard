#!/usr/bin/env python3
"""Site Closure Log generator with validation."""
import argparse, json, sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
import openpyxl

class ClosureGenerator:
    def __init__(self, workbook_path: str, project_code: str, closure_date: str):
        self.workbook_path = workbook_path
        self.project_code = project_code
        self.closure_date = datetime.strptime(closure_date, "%Y-%m-%d").date()
        self.warnings: List[str] = []
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True)

    def get_project_info(self) -> Optional[Dict[str, Any]]:
        if 'Projects' not in self.wb.sheetnames:
            return None
        sheet = self.wb['Projects']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0] == self.project_code:
                return {
                    'code': row[0],
                    'name': row[1],
                    'address': row[2],
                    'site_type': 'Warehouse with art storage'
                }
        return None

    def get_employee_info(self, employee_id: str) -> Optional[Dict[str, Any]]:
        if 'Employees' not in self.wb.sheetnames:
            return None
        sheet = self.wb['Employees']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0] == employee_id:
                return {
                    'id': row[0],
                    'name': row[1],
                    'trade': row[2] if len(row) > 2 else 'Foreman'
                }
        return None

    def get_closure_record(self) -> Optional[Dict[str, Any]]:
        if 'Site Closure Log' not in self.wb.sheetnames:
            self.warnings.append("VAL-001: Site Closure Log sheet not found")
            return None

        sheet = self.wb['Site Closure Log']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[1] and row[2]:
                row_date = row[1]
                if isinstance(row_date, str):
                    try:
                        row_date = datetime.strptime(row_date, "%Y-%m-%d").date()
                    except:
                        continue
                if row_date == self.closure_date and row[2] == self.project_code:
                    return {
                        'closure_id': row[0],
                        'date': str(row[1]),
                        'project_code': row[2],
                        'foreman_id': row[3],
                        'foreman_name': row[4],
                        'time_of_close': row[5],
                        'weather_at_close': row[6],
                        'equipment_left_overnight': row[7],
                        'checklist_values': row[8:32],
                        'notes': row[32],
                        'signed_timestamp': row[33]
                    }

        self.warnings.append("VAL-001: No closure record found for this date and project")
        return None

    def generate_json(self) -> Optional[Dict[str, Any]]:
        closure_rec = self.get_closure_record()
        if not closure_rec:
            return None

        project = self.get_project_info()
        if not project:
            return {'error': 'Project not found'}

        foreman_id = closure_rec['foreman_id']
        foreman_info = self.get_employee_info(foreman_id)
        if not foreman_info:
            self.warnings.append(f"VAL-002: Foreman ID {foreman_id} not found in Employees sheet")
            foreman_info = {'id': foreman_id, 'name': closure_rec['foreman_name'], 'trade': 'Foreman'}

        checklist_sections = [
            ('Personnel', [
                'All workers signed out',
                'Visitors escorted off + signed out',
                'No personnel remaining in building'
            ]),
            ('Equipment & Tools', [
                'Tools and equipment secured / removed',
                'Scaffold locked, tagged, retracted',
                'Compressors / generators shut off and secured',
                'Mast climbers lowered to ground and locked'
            ]),
            ('Hot Work / Fire Prevention', [
                '30-min fire watch completed for all hot work',
                'All heat sources confirmed cool',
                'Fire extinguishers in service position'
            ]),
            ('Dust & Debris Control', [
                'Dust collection emptied and sealed',
                'Path to interior art storage confirmed sealed',
                'Tarps / barriers secure for overnight'
            ]),
            ('Water Intrusion', [
                'Wall penetrations covered or sealed',
                'Window / opening protection in place'
            ]),
            ('Site Security', [
                'All site access points locked',
                'Construction fence secured',
                'Sidewalk shed pedestrian path clear'
            ]),
            ('Building Integrity', [
                'Building exterior doors locked',
                'Roof access closed and locked'
            ]),
            ('Climate / Interior Coordination', [
                'Building HVAC undisturbed',
                'Interior doors to art storage confirmed sealed'
            ]),
            ('Documentation', [
                'Daily report submitted',
                'Photos of site condition taken'
            ])
        ]

        checklist_values = closure_rec['checklist_values']
        sections = []
        total_items = 0
        completed_items = 0
        na_items = 0
        incomplete_items = 0
        idx = 0

        for section_name, items in checklist_sections:
            section_items = []
            for item in items:
                val = checklist_values[idx] if idx < len(checklist_values) else None
                status = 'na'
                if val == 'Y':
                    status = 'completed'
                    completed_items += 1
                elif val == 'N':
                    status = 'incomplete'
                    incomplete_items += 1
                    self.warnings.append(f"CLS-INCOMPLETE: {section_name}: {item}")
                elif val == 'N/A':
                    status = 'na'
                    na_items += 1

                section_items.append({'label': item, 'status': status})
                total_items += 1
                idx += 1

            sections.append({'name': section_name, 'items': section_items})

        completion_pct = 100 if (total_items - na_items) == 0 else int(100 * completed_items / (total_items - na_items))

        return {
            'closure_id': closure_rec['closure_id'],
            'project': {
                'code': project['code'],
                'name': project['name'],
                'address': project['address'],
                'site_type': project['site_type']
            },
            'closure': {
                'date': closure_rec['date'],
                'day_of_week': self._get_day_of_week(self.closure_date),
                'time_of_close': closure_rec['time_of_close'],
                'weather_at_close': closure_rec['weather_at_close'],
                'equipment_left_overnight': closure_rec['equipment_left_overnight'],
                'foreman': {
                    'employee_id': foreman_id,
                    'name': foreman_info['name'],
                    'trade': foreman_info['trade']
                }
            },
            'checklist': {
                'sections': sections,
                'summary': {
                    'total_items': total_items,
                    'completed_items': completed_items,
                    'na_items': na_items,
                    'incomplete_items': incomplete_items,
                    'completion_pct': completion_pct
                }
            },
            'notes': closure_rec['notes'],
            'signed_timestamp': closure_rec['signed_timestamp'],
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'source_workbook': str(Path(self.workbook_path).name),
                'warnings': self.warnings
            }
        }

    def _get_day_of_week(self, date_obj) -> str:
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        return days[date_obj.weekday()]

def main():
    parser = argparse.ArgumentParser(description='Generate Site Closure JSON')
    parser.add_argument('--workbook', required=True, help='Path to workbook')
    parser.add_argument('--project_code', required=True, help='Project code')
    parser.add_argument('--date', required=True, help='Closure date (YYYY-MM-DD)')
    parser.add_argument('--output_json', required=True, help='Output JSON path')

    args = parser.parse_args()

    gen = ClosureGenerator(args.workbook, args.project_code, args.date)
    data = gen.generate_json()

    if data is None:
        print(f"Error: Failed to generate closure", file=sys.stderr)
        sys.exit(1)

    Path(args.output_json).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output_json, 'w') as f:
        json.dump(data, f, indent=2)

    print(f"Generated: {args.output_json}")

if __name__ == '__main__':
    main()
